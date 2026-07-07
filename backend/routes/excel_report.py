from fastapi import APIRouter, Query, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from datetime import datetime

from main import get_current_user
from supabase_client import supabase

router = APIRouter()


@router.get("/reports/tasks/excel")
def generate_task_report(
    start_date: str = Query(..., description="Start date in YYYY-MM-DD format"),
    end_date: str = Query(..., description="End date in YYYY-MM-DD format"),
    status: str | None = Query(None),
    priority: str | None = Query(None),
    user: dict = Depends(get_current_user),
):
    # Only admins can generate workspace-wide reports
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Only admins can generate task reports.")

    workspace_id = user.get("workspace_id")
    if not workspace_id:
        raise HTTPException(status_code=400, detail="No workspace associated with this account.")

    # Validate date format early with a clear error, rather than a confusing query failure later
    try:
        datetime.strptime(start_date, "%Y-%m-%d")
        datetime.strptime(end_date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="Dates must be in YYYY-MM-DD format.")

    query = (
        supabase.table("tasks")
        .select("*")
        .eq("workspace_id", workspace_id)
        .gte("deadline", start_date)
        .lte("deadline", end_date)
    )

    if status:
        query = query.eq("status", status)
    if priority:
        query = query.eq("priority", priority)

    result = query.execute()
    tasks = result.data or []

    if not tasks:
        raise HTTPException(
            status_code=404,
            detail="No tasks found for the given filters and date range.",
        )

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Task Report"

    NAVY = "0B1B3D"

    headers = ["Task ID", "Title", "Assigned To", "Status", "Priority", "Deadline"]
    sheet.append(headers)

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    cell_border = Border(
        left=Side(style="thin", color="57595C"),
        right=Side(style="thin", color="57595C"),
        top=Side(style="thin", color="57595C"),
        bottom=Side(style="thin", color="57595C"),
    )

    for col_idx, _ in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = cell_border

    sheet.row_dimensions[1].height = 26

    for row_idx, task in enumerate(tasks, start=2):
        row_values = [
            task.get("id", ""),
            task.get("title", ""),
            task.get("assigned_to", ""),
            task.get("status", ""),
            task.get("priority", ""),
            task.get("deadline", ""),
        ]
        sheet.append(row_values)
        for col_idx in range(1, len(headers) + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.border = cell_border
            cell.alignment = Alignment(vertical="center")
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F4F6FA", end_color="F4F6FA", fill_type="solid")

    last_row = sheet.max_row
    last_col_letter = sheet.cell(row=1, column=len(headers)).column_letter
    table_range = f"A1:{last_col_letter}{max(last_row, 1)}"

    table = Table(displayName="TaskTable", ref=table_range)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=False
    )
    sheet.add_table(table)

    column_widths = [10, 28, 22, 14, 12, 14]
    for col_idx, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=col_idx).column_letter].width = width

    sheet.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"task_report_{start_date}_to_{end_date}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )